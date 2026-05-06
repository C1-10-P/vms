from typing import Dict
import pandas as pd
import openpyxl

from .models import AttendanceSession
import uuid
from django.utils import timezone
from django.core.cache import cache
from django.db import transaction
from django.db.models import Count, Q, Avg
from datetime import datetime, timedelta

import logging
logger = logging.getLogger(__name__)

import json

from apps.core.models import Student, Class
from apps.firmware.models import EdgeNode
from .models import ClassAttendance, DailyAttendanceSummary, VerificationLog
from apps.settings.models import AuditLog




class AttendanceService:
    """
    Business logic for attendance operations
    """
    
    @staticmethod
    def process_check_in(attendance):
        """
        Process student check-in with business rules
        Returns dict with success status and data
        """
        try:
            with transaction.atomic():
                # Check if already checked in today
                existing = ClassAttendance.objects.filter(
                    student=attendance.student,
                    class_obj=attendance.class_obj,
                    scan_time__date=timezone.now().date(),
                    verification_status='success'
                ).exists()
                
                if existing:
                    return {
                        'success': False,
                        'error': 'Student already checked in for this class today'
                    }
                
                # Check if class is currently in session
                current_time = timezone.now().time()
                schedule = attendance.class_obj.schedule
                weekday = timezone.now().strftime('%A').lower()
                
                if weekday in schedule:
                    session_start = datetime.strptime(schedule[weekday]['start'], '%H:%M').time()
                    session_end = datetime.strptime(schedule[weekday]['end'], '%H:%M').time()
                    
                    if current_time < session_start:
                        return {
                            'success': False,
                            'error': f'Class starts at {session_start}. Please wait.'
                        }
                    
                    # Mark as late if after start time + grace period
                    grace_minutes = 15
                    late_threshold = datetime.combine(
                        timezone.now().date(), session_start
                    ) + timedelta(minutes=grace_minutes)
                    
                    if timezone.now() > late_threshold:
                        attendance.verification_status = 'late'
                
                # Save attendance
                attendance.save()
                
                # Update daily summary
                AttendanceService._update_daily_summary(
                    attendance.class_obj, 
                    attendance.scan_time.date()
                )
                
                # Update cache
                cache_key = f"attendance_today_{attendance.student.id}"
                cache.delete(cache_key)
                
                # Create verification log
                VerificationLog.objects.create(
                    attendance=attendance,
                    student=attendance.student,
                    node=attendance.node,
                    event_type='success',
                    method=attendance.verification_method,
                    success=True,
                    processing_time_ms=0
                )
                
                # Trigger WebSocket notification
                AttendanceService._notify_realtime(attendance)
                
                return {
                    'success': True,
                    'attendance_id': attendance.id,
                    'message': 'Check-in successful'
                }
                
        except Exception as e:
            logger.error(f"Check-in failed: {e}")
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def process_api_check_in(student_id, class_code, node_uuid):
        """
        Process check-in from ESP32 device
        """
        try:
            # Look up student
            student = Student.objects.select_related('person').get(
                student_reg_number=student_id,
                is_active=True
            )
            
            # Look up class
            class_obj = Class.objects.get(
                class_code=class_code,
                is_active=True
            )
            
            # Look up node
            node = EdgeNode.objects.get(node_uuid=node_uuid)
            
            # Create attendance record
            attendance = ClassAttendance(
                student=student,
                class_obj=class_obj,
                node=node,
                scan_time=timezone.now(),
                verification_method='qr',
                verification_status='success'
            )
            
            result = AttendanceService.process_check_in(attendance)
            
            return {
                'success': result['success'],
                'student_name': student.person.full_name,
                'class_code': class_code,
                'timestamp': timezone.now().isoformat()
            }
            
        except Student.DoesNotExist:
            return {'success': False, 'error': 'Student not found'}
        except Class.DoesNotExist:
            return {'success': False, 'error': 'Class not found'}
        except EdgeNode.DoesNotExist:
            return {'success': False, 'error': 'Node not found'}
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def generate_report(start_date=None, end_date=None, department_id=None, class_id=None):
        """
        Generate attendance report with filters
        """
        queryset = ClassAttendance.objects.filter(verification_status='success')
        
        if start_date:
            queryset = queryset.filter(scan_time__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(scan_time__date__lte=end_date)
        
        if department_id:
            queryset = queryset.filter(student__department_id=department_id)
        
        if class_id:
            queryset = queryset.filter(class_obj_id=class_id)
        
        # Calculate statistics
        total_students = queryset.values('student_id').distinct().count()
        total_attendance = queryset.count()
        
        # Daily breakdown
        daily_stats = queryset.extra(
            {'date': "DATE(scan_time)"}
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        # Class breakdown
        class_stats = queryset.values(
            'class_obj__class_code',
            'class_obj__academic_unit__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Hourly distribution
        hourly_stats = queryset.extra(
            {'hour': "EXTRACT(HOUR FROM scan_time)"}
        ).values('hour').annotate(
            count=Count('id')
        ).order_by('hour')
        
        return {
            'start_date': start_date,
            'end_date': end_date,
            'total_students': total_students,
            'total_attendance': total_attendance,
            'average_daily': total_attendance / max(len(daily_stats), 1),
            'daily_stats': list(daily_stats),
            'class_stats': list(class_stats),
            'hourly_stats': list(hourly_stats)
        }
    
    @staticmethod
    def get_student_attendance_summary(student_id, days=30):
        """
        Get attendance summary for a specific student
        """
        cutoff_date = timezone.now() - timedelta(days=days)
        
        attendances = ClassAttendance.objects.filter(
            student_id=student_id,
            verification_status='success',
            scan_time__gte=cutoff_date
        ).select_related('class_obj__academic_unit')
        
        total_classes = attendances.count()
        unique_courses = attendances.values('class_obj__academic_unit__code').distinct().count()
        
        # Attendance by course
        by_course = attendances.values(
            'class_obj__academic_unit__code',
            'class_obj__academic_unit__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Daily attendance
        daily = attendances.extra(
            {'date': "DATE(scan_time)"}
        ).values('date').annotate(
            count=Count('id')
        ).order_by('-date')
        
        return {
            'student_id': student_id,
            'period_days': days,
            'total_attendances': total_classes,
            'unique_courses': unique_courses,
            'by_course': list(by_course),
            'daily_attendance': list(daily),
            'attendance_rate': round((total_classes / days) * 100, 2) if days > 0 else 0
        }
    
    @staticmethod
    def _update_daily_summary(class_obj, date):
        """
        Update daily attendance summary for a class
        """
        summary, created = DailyAttendanceSummary.objects.get_or_create(
            class_obj=class_obj,
            summary_date=date
        )
        
        total_attendance = ClassAttendance.objects.filter(
            class_obj=class_obj,
            scan_time__date=date,
            verification_status='success'
        ).count()
        
        summary.present_count = total_attendance
        summary.total_students = class_obj.enrolled_count
        summary.attendance_percentage = (total_attendance / class_obj.enrolled_count * 100) if class_obj.enrolled_count > 0 else 0
        summary.save()
        
        return summary
    
    @staticmethod
    def _notify_realtime(attendance):
        """
        Send WebSocket notification for real-time updates
        """
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                'attendance_live',
                {
                    'type': 'attendance_update',
                    'data': {
                        'student_id': attendance.student.student_reg_number,
                        'student_name': attendance.student.person.full_name,
                        'class_code': attendance.class_obj.class_code,
                        'timestamp': attendance.scan_time.isoformat()
                    }
                }
            )
        except Exception as e:
            logger.error(f"WebSocket notification failed: {e}")
    
    @staticmethod
    def bulk_attendance_checkin(records):
        """
        Process multiple attendance records in bulk
        """
        results = {
            'successful': [],
            'failed': [],
            'total': len(records)
        }
        
        with transaction.atomic():
            for record in records:
                try:
                    student = Student.objects.get(student_reg_number=record.get('student_id'))
                    class_obj = Class.objects.get(class_code=record.get('class_code'))
                    
                    attendance = ClassAttendance.objects.create(
                        student=student,
                        class_obj=class_obj,
                        scan_time=timezone.now(),
                        verification_method=record.get('method', 'bulk'),
                        verification_status='success'
                    )
                    
                    results['successful'].append({
                        'student_id': student.student_reg_number,
                        'attendance_id': attendance.id
                    })
                    
                except Exception as e:
                    results['failed'].append({
                        'student_id': record.get('student_id'),
                        'error': str(e)
                    })
        
        return results
    
    @staticmethod
    def get_attendance_analytics(start_date, end_date):
        """
        Get advanced attendance analytics
        """
        queryset = ClassAttendance.objects.filter(
            scan_time__date__gte=start_date,
            scan_time__date__lte=end_date,
            verification_status='success'
        )
        
        # Overall statistics
        total = queryset.count()
        unique_students = queryset.values('student').distinct().count()
        
        # Best and worst performing classes
        class_performance = queryset.values(
            'class_obj__class_code',
            'class_obj__academic_unit__name'
        ).annotate(
            attendance_count=Count('id'),
            unique_students=Count('student', distinct=True)
        ).order_by('-attendance_count')
        
        # Peak attendance hours
        peak_hours = queryset.extra(
            {'hour': "EXTRACT(HOUR FROM scan_time)"}
        ).values('hour').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return {
            'period': {
                'start': start_date,
                'end': end_date,
                'days': (end_date - start_date).days + 1
            },
            'total_attendance': total,
            'unique_students': unique_students,
            'average_daily': total / max((end_date - start_date).days, 1),
            'class_performance': list(class_performance[:10]),
            'peak_hours': list(peak_hours)
        }
    

class AttendanceSessionService:
    """
    Service for managing attendance sessions
    """
    
    @staticmethod
    def create_session(student_reg: str = None, class_code: str = None, 
                       scan_method: str = 'qr', scan_device: str = None) -> Dict:
        """
        Create a new attendance session
        """
        from .models import AttendanceSession
        
        session_id = str(uuid.uuid4())
        expires_at = timezone.now() + timedelta(minutes=5)  # 5 minute window
        
        session = AttendanceSession.objects.create(
            session_id=session_id,
            student_reg_number=student_reg,
            class_code=class_code,
            scan_method=scan_method,
            scan_device=scan_device,
            expires_at=expires_at,
            status='pending'
        )
        
        return {
            'success': True,
            'session_id': session_id,
            'expires_at': expires_at.isoformat(),
            'message': 'Session created. Please validate within 5 minutes.'
        }
    
   
    @staticmethod
    # services.py
    @staticmethod
    def validate_session(session_id):
        session = AttendanceSession.objects.get(session_id=session_id)
        
        # Update fields directly (don't call session.validate() to avoid recursion!)
        session.status = 'validated'
        session.save()
        
        # Log the action (Fixes the Audit Log error)
        try:
            AuditLog.objects.create(session=session, action="VALIDATION")
        except Exception as e:
            print(f"Failed to create audit log: {e}")
            
        return session # Returns an object, so update the view as shown in Step 1
    
    @staticmethod
    def process_full_attendance_flow(student_reg: str, class_code: str = None,
                                      scan_method: str = 'qr', scan_device: str = None) -> Dict:
        """
        Complete flow: Create session → Validate → Create attendance
        """
        # Step 1: Create session
        session_result = AttendanceSessionService.create_session(
            student_reg=student_reg,
            class_code=class_code,
            scan_method=scan_method,
            scan_device=scan_device
        )
        
        if not session_result['success']:
            return session_result
        
        # Step 2: Validate session
        validation_result = AttendanceSessionService.validate_session(
            session_result['session_id']
        )
        
        return {
            'session': session_result,
            'attendance': validation_result
        }


class AttendanceImportService:
    """
    Service for importing attendance records from Excel files
    Supports multiple formats: .xlsx, .xls, .csv
    """
    
    # Expected column headers
    REQUIRED_COLUMNS = ['student_reg', 'class_code', 'date', 'status']
    OPTIONAL_COLUMNS = ['time', 'method', 'remarks']
    
    # Status mapping
    STATUS_MAPPING = {
        'present': 'success',
        'absent': 'failed',
        'late': 'late',
        'excused': 'excused',
        'P': 'success',
        'A': 'failed',
        'L': 'late',
        'E': 'excused',
        1: 'success',
        0: 'failed'
    }
    
    def __init__(self):
        self.imported_count = 0
        self.skipped_count = 0
        self.errors = []
    
    def import_from_excel(self, excel_file, user=None):
        """
        Import attendance records from Excel file
        Returns dict with success status and counts
        """
        try:
            # Read the Excel file
            if excel_file.name.endswith('.csv'):
                df = pd.read_csv(excel_file)
            else:
                df = pd.read_excel(excel_file, engine='openpyxl')
            
            # Validate required columns
            missing_cols = [col for col in self.REQUIRED_COLUMNS if col not in df.columns]
            if missing_cols:
                return {
                    'success': False,
                    'error': f'Missing required columns: {", ".join(missing_cols)}'
                }
            
            # Clean and process data
            df = self._clean_dataframe(df)
            
            # Process each row
            with transaction.atomic():
                for index, row in df.iterrows():
                    self._process_row(row, index + 2, user)  # +2 for 1-indexed and header row
            
            return {
                'success': True,
                'imported': self.imported_count,
                'skipped': self.skipped_count,
                'errors': self.errors[:10]  # Return first 10 errors
            }
            
        except Exception as e:
            logger.error(f"Excel import failed: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _clean_dataframe(self, df):
        """Clean and prepare dataframe"""
        # Strip whitespace from column names
        df.columns = df.columns.str.strip().str.lower()
        
        # Strip whitespace from string columns
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].astype(str).str.strip()
        
        # Convert date column to datetime
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
        
        # Convert time column if exists
        if 'time' in df.columns:
            df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S', errors='coerce').dt.time
        
        # Map status values
        if 'status' in df.columns:
            df['status'] = df['status'].map(lambda x: self.STATUS_MAPPING.get(str(x).lower(), 'failed'))
        
        # Drop rows with missing required data
        df = df.dropna(subset=['student_reg', 'class_code', 'date', 'status'])
        
        return df
    
    def _process_row(self, row, row_number, user):
        """Process a single row of attendance data"""
        try:
            # Look up student
            student = Student.objects.filter(
                student_reg_number=row['student_reg'],
                is_active=True
            ).first()
            
            if not student:
                self.skipped_count += 1
                self.errors.append(f"Row {row_number}: Student '{row['student_reg']}' not found")
                return
            
            # Look up class
            class_obj = Class.objects.filter(
                class_code=row['class_code'],
                is_active=True
            ).first()
            
            if not class_obj:
                self.skipped_count += 1
                self.errors.append(f"Row {row_number}: Class '{row['class_code']}' not found")
                return
            
            # Build scan time
            scan_date = row['date']
            scan_time = row.get('time', datetime.now().time())
            scan_datetime = datetime.combine(scan_date.date(), scan_time) if hasattr(scan_date, 'date') else scan_date
            
            if timezone.is_naive(scan_datetime):
                scan_datetime = timezone.make_aware(scan_datetime)
            
            # Check for duplicate
            existing = ClassAttendance.objects.filter(
                student=student,
                class_obj=class_obj,
                scan_time__date=scan_datetime.date()
            ).exists()
            
            if existing:
                self.skipped_count += 1
                self.errors.append(f"Row {row_number}: Duplicate attendance for {student.student_reg_number} on {scan_datetime.date()}")
                return
            
            # Create attendance record
            verification_method = row.get('method', 'import')
            if verification_method not in ['qr', 'face', 'rfid', 'manual', 'import']:
                verification_method = 'import'
            
            attendance = ClassAttendance.objects.create(
                student=student,
                class_obj=class_obj,
                scan_time=scan_datetime,
                verification_method=verification_method,
                verification_status=row['status'],
                remarks=row.get('remarks', f'Imported from Excel - Row {row_number}')
            )
            
            self.imported_count += 1
            
        except Exception as e:
            self.skipped_count += 1
            self.errors.append(f"Row {row_number}: {str(e)}")
    
    @staticmethod
    def download_template():
        """
        Generate Excel template for attendance import
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Template"
        
        # Define headers
        headers = ['student_reg', 'class_code', 'date', 'status', 'time', 'method', 'remarks']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0A6232", end_color="0A6232", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add example data
        examples = [
            ['ENE221-0108/2018', 'TIE4101', '2024-01-15', 'present', '09:30:00', 'qr', ''],
            ['ENE221-0100/2020', 'TIE4101', '2024-01-15', 'absent', '', 'manual', 'Sick'],
            ['SCT211-001/2021', 'TIE4101', '2024-01-15', 'late', '09:45:00', 'face', 'Traffic'],
        ]
        
        for row_idx, example in enumerate(examples, 2):
            for col_idx, value in enumerate(example, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add notes
        notes_row = len(examples) + 3
        ws.cell(row=notes_row, column=1, value="Instructions:")
        ws.cell(row=notes_row + 1, column=1, value="1. student_reg: Student registration number (required)")
        ws.cell(row=notes_row + 2, column=1, value="2. class_code: Class code (required)")
        ws.cell(row=notes_row + 3, column=1, value="3. date: Date of attendance in YYYY-MM-DD format (required)")
        ws.cell(row=notes_row + 4, column=1, value="4. status: present/absent/late/excused (required)")
        ws.cell(row=notes_row + 5, column=1, value="5. time: Time of attendance in HH:MM:SS format (optional)")
        ws.cell(row=notes_row + 6, column=1, value="6. method: qr/face/rfid/manual (optional, defaults to 'import')")
        ws.cell(row=notes_row + 7, column=1, value="7. remarks: Additional notes (optional)")
        
        # Adjust column widths
        for col in headers:
            column_letter = openpyxl.utils.get_column_letter(headers.index(col) + 1)
            ws.column_dimensions[column_letter].width = 20
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


class AttendanceExportService:
    """
    Service for exporting attendance records to Excel
    """
    
    @staticmethod
    def export_to_excel(attendances):
        """
        Export attendance records to Excel
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Records"
        
        # Define headers
        headers = [
            'Date', 'Time', 'Student Registration', 'Student Name', 
            'Class Code', 'Course', 'Status', 'Method', 'Remarks'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0A6232", end_color="0A6232", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, attendance in enumerate(attendances, 2):
            # Status color
            status_color = {
                'success': '92D050',  # Green
                'late': 'FFC000',      # Orange
                'failed': 'FF6B6B',    # Red
                'excused': '5B9BD5'    # Blue
            }.get(attendance.verification_status, 'FFFFFF')
            
            ws.cell(row=row_idx, column=1, value=attendance.scan_time.strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=2, value=attendance.scan_time.strftime('%H:%M:%S'))
            ws.cell(row=row_idx, column=3, value=attendance.student.student_reg_number)
            ws.cell(row=row_idx, column=4, value=attendance.student.person.full_name)
            ws.cell(row=row_idx, column=5, value=attendance.class_obj.class_code)
            ws.cell(row=row_idx, column=6, value=attendance.class_obj.academic_unit.name)
            ws.cell(row=row_idx, column=7, value=attendance.get_verification_status_display())
            ws.cell(row=row_idx, column=8, value=attendance.get_verification_method_display())
            ws.cell(row=row_idx, column=9, value=attendance.remarks or '')
            
            # Apply border and status color
            for col in range(1, 10):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                if col == 7:  # Status column
                    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
        
        # Adjust column widths
        column_widths = [12, 10, 20, 25, 15, 30, 12, 12, 30]
        for i, width in enumerate(column_widths, 1):
            column_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[column_letter].width = width
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="Attendance Summary Report")
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        summary_data = [
            ("Total Records", len(attendances)),
            ("Present", sum(1 for a in attendances if a.verification_status == 'success')),
            ("Absent", sum(1 for a in attendances if a.verification_status == 'failed')),
            ("Late", sum(1 for a in attendances if a.verification_status == 'late')),
            ("Excused", sum(1 for a in attendances if a.verification_status == 'excused')),
            ("", ""),
            ("Report Generated", timezone.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 3):
            ws_summary.cell(row=row_idx, column=1, value=label)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


class AttendanceBulkService:
    """
    Service for bulk attendance operations
    """
    
    @staticmethod
    def bulk_mark_attendance(class_id, date, student_statuses, user=None):
        """
        Bulk mark attendance for a class on a given date
        student_statuses: dict of {student_id: status}
        """
        from apps.classroom.models import ClassAttendance
        
        class_obj = Class.objects.filter(id=class_id, is_active=True).first()
        if not class_obj:
            return {'success': False, 'error': 'Class not found'}
        
        created = 0
        updated = 0
        errors = []
        
        with transaction.atomic():
            for student_id, status in student_statuses.items():
                try:
                    student = Student.objects.get(id=student_id, is_active=True)
                    
                    # Check if attendance already exists for this date
                    attendance, created_flag = ClassAttendance.objects.update_or_create(
                        student=student,
                        class_obj=class_obj,
                        scan_time__date=date,
                        defaults={
                            'scan_time': timezone.make_aware(datetime.combine(date, datetime.now().time())),
                            'verification_status': status,
                            'verification_method': 'bulk',
                            'remarks': f'Bulk update by {user.username if user else "System"}'
                        }
                    )
                    
                    if created_flag:
                        created += 1
                    else:
                        updated += 1
                        
                except Exception as e:
                    errors.append(f"Student {student_id}: {str(e)}")
        
        return {
            'success': True,
            'created': created,
            'updated': updated,
            'errors': errors
        }