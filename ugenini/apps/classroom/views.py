from django.shortcuts import redirect, render, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Avg, Sum
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import csv
import json
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.db import transaction

from apps.users.decorators import permission_required, api_permission_required
from apps.users.permissions import VMSPermissions, PermissionChecker
from apps.core.models.department import Department
from .models import ClassAttendance, DailyAttendanceSummary, VerificationLog
from .services import AttendanceService
from apps.core.models import Student, Class

@method_decorator(permission_required(VMSPermissions.ATTENDANCE_CREATE), name='dispatch')
class AttendanceCheckInView(LoginRequiredMixin, CreateView):
    """
    Manual attendance check-in view
    """
    model = ClassAttendance
    fields = ['student', 'class_obj']
    template_name = 'classroom/check_in.html'
    success_url = reverse_lazy('classroom:check_in')
    
    
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get today's date
        today = timezone.now().date()
        
        # Get classes for dropdown (only active classes)
        context['classes'] = Class.objects.filter(
            is_active=True,
            start_date__lte=today,
            end_date__gte=today
        ).select_related('academic_unit')
        
        # Get recent attendance
        context['recent_attendances'] = ClassAttendance.objects.filter(
            scan_time__date=today
        ).select_related(
            'student__person', 'class_obj'
        ).order_by('-scan_time')[:20]
        
        # Get statistics
        context['today_count'] = ClassAttendance.objects.filter(
            scan_time__date=today,
            verification_status='success'
        ).count()
        
        context['unique_students'] = ClassAttendance.objects.filter(
            scan_time__date=today,
            verification_status='success'
        ).values('student').distinct().count()
        
        return context
    
    def form_valid(self, form):
        attendance = form.save(commit=False)
        attendance.scan_time = timezone.now()
        attendance.verification_method = 'manual'
        attendance.verification_status = 'success'
        
        if hasattr(self.request.user, 'person') and hasattr(self.request.user.person, 'staff'):
            attendance.recorded_by = self.request.user.person.staff
        
        # Process check-in
        service = AttendanceService()
        result = service.process_check_in(attendance)
        
        if result['success']:
            messages.success(
                self.request,
                f"Check-in successful for {attendance.student.person.full_name}"
            )
        else:
            messages.error(self.request, result['error'])
            return self.form_invalid(form)
        
        return super().form_valid(form)


class AttendanceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    List view for attendance records
    """
    model = ClassAttendance
    ordering = ['-created_at']
    template_name = 'classroom/list.html'
    context_object_name = 'attendances'
    paginate_by = 50
    
    # This replaces the @permission_required decorator
    permission_required = VMSPermissions.ATTENDANCE_VIEW

    def get_queryset(self):
        # We start with the base queryset
        queryset = super().get_queryset().select_related(
            'student__person', 'class_obj__academic_unit', 'node'
        )
        
        # Date range filters
        start_date = self.request.GET.get('start_date')
        if start_date:
            queryset = queryset.filter(scan_time__date__gte=start_date)
        
        end_date = self.request.GET.get('end_date')
        if end_date:
            queryset = queryset.filter(scan_time__date__lte=end_date)
        
        # Filter by student
        student_id = self.request.GET.get('student_id')
        if student_id:
            queryset = queryset.filter(student_id=student_id)
        
        # Filter by class
        class_id = self.request.GET.get('class_id')
        if class_id:
            queryset = queryset.filter(class_obj_id=class_id)
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(verification_status=status)
        
        # Filter by department (if user has limited access)
        if not PermissionChecker.has_permission(self.request.user, VMSPermissions.ATTENDANCE_VIEW):
            if hasattr(self.request.user, 'person') and hasattr(self.request.user.person, 'staff'):
                department_id = self.request.user.person.staff.department_id
                queryset = queryset.filter(class_obj__department_id=department_id)
        
        # Apply ordering
        order_by = self.request.GET.get('order_by', '-scan_time')
        if order_by:
            queryset = queryset.order_by(order_by)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters to persist in search forms
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['student_id'] = self.request.GET.get('student_id', '')
        context['class_id'] = self.request.GET.get('class_id', '')
        context['status'] = self.request.GET.get('status', '')
        context['order_by'] = self.request.GET.get('order_by', '-scan_time')
        
        # Get statistics for the ALREADY filtered queryset
        filtered_qs = self.get_queryset()
        context['total_count'] = filtered_qs.count()
        context['success_count'] = filtered_qs.filter(verification_status='success').count()
        context['failed_count'] = filtered_qs.filter(verification_status='failed').count()
        context['unique_students'] = filtered_qs.values('student').distinct().count()
        
        # Get classes for filter dropdown
        context['classes'] = Class.objects.filter(is_active=True).select_related('academic_unit')
        
        return context


class AttendanceReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """
    Attendance report generation view
    """
    template_name = 'classroom/report.html'
    
    # 1. Standard Django way to enforce permissions in CBVs
    permission_required = VMSPermissions.REPORT_GENERATE
    
    # 2. Removed the @permission_required decorator from dispatch entirely.
    # The Mixin above handles the check automatically.
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get report parameters
        report_type = self.request.GET.get('report_type', 'daily')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        department_id = self.request.GET.get('department_id')
        class_id = self.request.GET.get('class_id')
        
        # Generate report based on type
        if report_type == 'daily':
            context['report'] = self.generate_daily_report(start_date, end_date, department_id, class_id)
        elif report_type == 'weekly':
            context['report'] = self.generate_weekly_report(start_date, end_date, department_id, class_id)
        elif report_type == 'monthly':
            context['report'] = self.generate_monthly_report(start_date, end_date, department_id, class_id)
        elif report_type == 'student':
            context['report'] = self.generate_student_report(self.request.GET.get('student_id'))
        
        context['report_type'] = report_type
        context['departments'] = Department.objects.filter(is_active=True)
        context['classes'] = Class.objects.filter(is_active=True)
        
        return context
    
    def generate_daily_report(self, start_date, end_date, department_id, class_id):
        """Generate daily attendance report"""
        from apps.classroom.models import ClassAttendance
        from django.db.models.functions import TruncDate
        
        queryset = ClassAttendance.objects.filter(verification_status='success')
        
        if start_date:
            queryset = queryset.filter(scan_time__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(scan_time__date__lte=end_date)
        if department_id:
            queryset = queryset.filter(student__department_id=department_id)
        if class_id:
            queryset = queryset.filter(class_obj_id=class_id)
        
        # Group by date
        daily_stats = queryset.annotate(
            date=TruncDate('scan_time')
        ).values('date').annotate(
            count=Count('id'),
            unique_students=Count('student', distinct=True)
        ).order_by('date')
        
        return {
            'type': 'daily',
            'start_date': start_date,
            'end_date': end_date,
            'data': list(daily_stats),
            'total_attendance': sum(item['count'] for item in daily_stats),
            'average_daily': sum(item['count'] for item in daily_stats) / max(len(daily_stats), 1)
        }
    
    def generate_weekly_report(self, start_date, end_date, department_id, class_id):
        """Generate weekly attendance report"""
        from apps.classroom.models import ClassAttendance
        
        queryset = ClassAttendance.objects.filter(verification_status='success')
        
        if start_date:
            queryset = queryset.filter(scan_time__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(scan_time__date__lte=end_date)
        if department_id:
            queryset = queryset.filter(student__department_id=department_id)
        if class_id:
            queryset = queryset.filter(class_obj_id=class_id)
        
        # Get weeks
        weeks = {}
        for attendance in queryset:
            week_number = attendance.scan_time.isocalendar()[1]
            week_key = f"Week {week_number}"
            if week_key not in weeks:
                weeks[week_key] = 0
            weeks[week_key] += 1
        
        return {
            'type': 'weekly',
            'data': [{'week': k, 'count': v} for k, v in weeks.items()],
            'total_attendance': sum(weeks.values())
        }
    
    def generate_monthly_report(self, start_date, end_date, department_id, class_id):
        """Generate monthly attendance report"""
        from apps.classroom.models import ClassAttendance
        
        queryset = ClassAttendance.objects.filter(verification_status='success')
        
        if start_date:
            queryset = queryset.filter(scan_time__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(scan_time__date__lte=end_date)
        if department_id:
            queryset = queryset.filter(student__department_id=department_id)
        if class_id:
            queryset = queryset.filter(class_obj_id=class_id)
        
        # Group by month
        monthly_stats = {}
        for attendance in queryset:
            month_key = attendance.scan_time.strftime('%Y-%m')
            if month_key not in monthly_stats:
                monthly_stats[month_key] = 0
            monthly_stats[month_key] += 1
        
        return {
            'type': 'monthly',
            'data': [{'month': k, 'count': v} for k, v in monthly_stats.items()],
            'total_attendance': sum(monthly_stats.values())
        }
    
    def generate_student_report(self, student_id):
        """Generate report for specific student"""
        from apps.classroom.models import ClassAttendance
        
        if not student_id:
            return None
        
        student = get_object_or_404(Student, id=student_id)
        
        # Get all attendance for this student
        attendances = ClassAttendance.objects.filter(
            student=student,
            verification_status='success'
        ).select_related('class_obj__academic_unit')
        
        # Calculate attendance percentage
        total_classes = Class.objects.filter(program=student.program).count()
        attended = attendances.count()
        
        # Group by course
        course_attendance = {}
        for att in attendances:
            course_code = att.class_obj.academic_unit.code
            if course_code not in course_attendance:
                course_attendance[course_code] = 0
            course_attendance[course_code] += 1
        
        return {
            'student': student,
            'total_attended': attended,
            'total_classes': total_classes,
            'percentage': (attended / total_classes * 100) if total_classes > 0 else 0,
            'course_breakdown': course_attendance,
            'recent_attendance': attendances[:20]
        }


@csrf_exempt
@api_permission_required(VMSPermissions.ATTENDANCE_CREATE)
def api_attendance_checkin(request):
    """
    API endpoint for ESP32 devices to submit attendance
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Parse request body
        data = json.loads(request.body)
        
        student_id = data.get('student_id')
        class_code = data.get('class_code')
        node_uuid = data.get('node_uuid')
        verification_method = data.get('method', 'qr')
        
        # Validate required fields
        if not student_id or not class_code:
            return JsonResponse({
                'error': 'Missing required fields: student_id, class_code'
            }, status=400)
        
        # Process check-in
        service = AttendanceService()
        result = service.process_api_check_in(
            student_id=student_id,
            class_code=class_code,
            node_uuid=node_uuid
        )
        
        if result['success']:
            return JsonResponse(result, status=200)
        else:
            return JsonResponse(result, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@permission_required(VMSPermissions.ATTENDANCE_EXPORT)
def export_attendance_csv(request):
    """
    Export attendance records to CSV
    """
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    class_id = request.GET.get('class_id')
    
    # Build queryset
    queryset = ClassAttendance.objects.select_related(
        'student__person', 'class_obj__academic_unit'
    )
    
    if start_date:
        queryset = queryset.filter(scan_time__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(scan_time__date__lte=end_date)
    if class_id:
        queryset = queryset.filter(class_obj_id=class_id)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Date', 'Time', 'Student ID', 'Student Name', 'Class Code',
        'Course', 'Verification Method', 'Status', 'IP Address'
    ])
    
    for attendance in queryset:
        writer.writerow([
            attendance.scan_time.date(),
            attendance.scan_time.time(),
            attendance.student.student_reg_number,
            attendance.student.person.full_name,
            attendance.class_obj.class_code,
            attendance.class_obj.academic_unit.name,
            attendance.get_verification_method_display(),
            attendance.get_verification_status_display(),
            attendance.ip_address or ''
        ])
    
    return response

@method_decorator(login_required, name='dispatch')
class AttendanceUpdateView(LoginRequiredMixin, UpdateView):
    model = ClassAttendance
    fields = ['verification_status', 'remarks']
    template_name = 'classroom/attendance_form.html'
    
    def get_success_url(self):
        return reverse_lazy('classroom:detail', kwargs={'pk': self.object.pk})

@method_decorator(login_required, name='dispatch')
class AttendanceVerifyView(LoginRequiredMixin, UpdateView):
    model = ClassAttendance
    fields = ['verification_status']
    template_name = 'classroom/verify.html'
    
    def form_valid(self, form):
        form.instance.verified_by = self.request.user.person.staff
        form.instance.verified_at = timezone.now()
        messages.success(self.request, 'Attendance verified successfully.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('classroom:detail', kwargs={'pk': self.object.pk})


@login_required
def daily_attendance_report(request):
    """Generate daily attendance report"""
    date = request.GET.get('date', timezone.now().date())
    attendances = ClassAttendance.objects.filter(
        scan_time__date=date,
        verification_status='success'
    ).select_related('student__person', 'class_obj')
    
    context = {
        'date': date,
        'attendances': attendances,
        'total_count': attendances.count(),
        'unique_students': attendances.values('student').distinct().count()
    }
    return render(request, 'classroom/daily_report.html', context)


@login_required
def weekly_attendance_report(request):
    """Generate weekly attendance report"""
    from datetime import timedelta
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=7)
    
    from django.db.models.functions import TruncDate
    from django.db.models import Count
    
    stats = ClassAttendance.objects.filter(
        scan_time__date__gte=start_date,
        scan_time__date__lte=end_date,
        verification_status='success'
    ).annotate(date=TruncDate('scan_time')).values('date').annotate(
        count=Count('id')
    ).order_by('date')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'stats': stats,
        'total': ClassAttendance.objects.filter(
            scan_time__date__gte=start_date,
            scan_time__date__lte=end_date
        ).count()
    }
    return render(request, 'classroom/weekly_report.html', context)


@login_required
def monthly_attendance_report(request):
    """Generate monthly attendance report"""
    month = request.GET.get('month', timezone.now().month)
    year = request.GET.get('year', timezone.now().year)
    
    stats = ClassAttendance.objects.filter(
        scan_time__year=year,
        scan_time__month=month,
        verification_status='success'
    )
    
    context = {
        'month': month,
        'year': year,
        'total': stats.count(),
        'unique_students': stats.values('student').distinct().count()
    }
    return render(request, 'classroom/monthly_report.html', context)


@login_required
def student_attendance_report(request, student_id):
    """Generate attendance report for a specific student"""
    student = get_object_or_404(Student, id=student_id)
    attendances = ClassAttendance.objects.filter(
        student=student,
        verification_status='success'
    ).select_related('class_obj')
    
    context = {
        'student': student,
        'attendances': attendances,
        'total': attendances.count()
    }
    return render(request, 'classroom/student_report.html', context)


@permission_required(VMSPermissions.ATTENDANCE_EXPORT)
def export_attendance_excel(request):
    """Export attendance to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"
    
    # Headers
    headers = ['Date', 'Time', 'Student ID', 'Student Name', 'Class', 'Status', 'Method']
    ws.append(headers)
    
    # Data
    attendances = ClassAttendance.objects.select_related('student__person', 'class_obj')
    for att in attendances:
        ws.append([
            att.scan_time.date(),
            att.scan_time.time(),
            att.student.student_reg_number,
            att.student.person.full_name,
            att.class_obj.class_code,
            att.get_verification_status_display(),
            att.get_verification_method_display()
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'
    wb.save(response)
    return response


@permission_required(VMSPermissions.ATTENDANCE_EXPORT)
def export_attendance_pdf(request):
    """Export attendance to PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    story.append(Paragraph("Attendance Report", styles['Title']))
    story.append(Spacer(1, 12))
    
    # Table data
    attendances = ClassAttendance.objects.select_related('student__person', 'class_obj')[:50]
    table_data = [['Date', 'Student Name', 'Class', 'Status']]
    for att in attendances:
        table_data.append([
            att.scan_time.strftime('%Y-%m-%d %H:%M'),
            att.student.person.full_name,
            att.class_obj.class_code,
            att.get_verification_status_display()
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    return response


@csrf_exempt
def api_verify_attendance(request):
    """API endpoint to verify attendance"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    data = json.loads(request.body)
    attendance_id = data.get('attendance_id')
    status = data.get('status')
    
    try:
        attendance = ClassAttendance.objects.get(id=attendance_id)
        attendance.verification_status = status
        attendance.save()
        return JsonResponse({'success': True})
    except ClassAttendance.DoesNotExist:
        return JsonResponse({'error': 'Attendance not found'}, status=404)


@csrf_exempt
def api_bulk_attendance(request):
    """Bulk attendance submission API"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    data = json.loads(request.body)
    records = data.get('records', [])
    
    created = []
    errors = []
    
    with transaction.atomic():
        for record in records:
            try:
                attendance = ClassAttendance.objects.create(
                    student_id=record.get('student_id'),
                    class_obj_id=record.get('class_id'),
                    scan_time=timezone.now(),
                    verification_method=record.get('method', 'api')
                )
                created.append(attendance.id)
            except Exception as e:
                errors.append({'record': record, 'error': str(e)})
    
    return JsonResponse({
        'success': True,
        'created': len(created),
        'errors': errors
    })


@login_required
def ajax_search_students(request):
    """AJAX endpoint for student search"""
    query = request.GET.get('q', '')
    students = Student.objects.filter(
        Q(student_reg_number__icontains=query) |
        Q(person__first_name__icontains=query) |
        Q(person__last_name__icontains=query),
        is_active=True
    ).select_related('person')[:10]
    
    results = [{
        'id': s.id,
        'reg_number': s.student_reg_number,
        'name': s.person.full_name
    } for s in students]
    
    return JsonResponse({'results': results})


@login_required
def ajax_get_class_schedule(request):
    """AJAX endpoint to get class schedule"""
    class_id = request.GET.get('class_id')
    class_obj = get_object_or_404(Class, id=class_id)
    return JsonResponse({'schedule': class_obj.schedule})


@login_required
def ajax_recent_activity(request):
    """AJAX endpoint for recent attendance activity"""
    attendances = ClassAttendance.objects.filter(
        scan_time__gte=timezone.now() - timezone.timedelta(minutes=30)
    ).select_related('student__person', 'class_obj')[:20]
    
    data = [{
        'student_name': a.student.person.full_name,
        'class_code': a.class_obj.class_code,
        'time': a.scan_time.strftime('%H:%M:%S'),
        'status': a.verification_status
    } for a in attendances]
    
    return JsonResponse({'activity': data})


@login_required
def attendance_summary(request):
    """Attendance summary statistics"""
    today = timezone.now().date()
    week_ago = today - timezone.timedelta(days=7)
    
    summary = {
        'today': ClassAttendance.objects.filter(scan_time__date=today).count(),
        'this_week': ClassAttendance.objects.filter(scan_time__date__gte=week_ago).count(),
        'total': ClassAttendance.objects.count(),
        'by_method': list(ClassAttendance.objects.values('verification_method').annotate(
            count=Count('id')
        ))
    }
    return JsonResponse(summary)


@login_required
def attendance_trends(request):
    """Attendance trend data for charts"""
    from datetime import timedelta
    
    trends = []
    for i in range(6, -1, -1):
        date = timezone.now().date() - timedelta(days=i)
        count = ClassAttendance.objects.filter(
            scan_time__date=date,
            verification_status='success'
        ).count()
        trends.append({'date': date.isoformat(), 'count': count})
    
    return JsonResponse({'trends': trends})


@login_required
def attendance_by_course(request):
    """Attendance breakdown by course"""
    stats = ClassAttendance.objects.filter(
        verification_status='success'
    ).values('class_obj__academic_unit__code', 'class_obj__academic_unit__name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    return JsonResponse({'by_course': list(stats)})

def api_students(request):
    """API endpoint for student search"""
    students = Student.objects.select_related('person').filter(is_active=True)[:100]
    data = []
    for student in students:
        data.append({
            'id': student.id,
            'name': student.person.full_name(),
            'reg_number': student.student_reg_number
        })
    return JsonResponse(data, safe=False)

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Q

@login_required
def export_attendance_excel(request):
    """Export attendance records to Excel"""
    # Build queryset with same filters as the list view
    queryset = ClassAttendance.objects.select_related(
        'student__person', 'class_obj__academic_unit'
    ).order_by('-scan_time')
    
    # Apply filters
    start_date = request.GET.get('start_date')
    if start_date:
        queryset = queryset.filter(scan_time__date__gte=start_date)
    end_date = request.GET.get('end_date')
    if end_date:
        queryset = queryset.filter(scan_time__date__lte=end_date)
    student_id = request.GET.get('student_id')
    if student_id:
        queryset = queryset.filter(student_id=student_id)
    class_id = request.GET.get('class_id')
    if class_id:
        queryset = queryset.filter(class_obj_id=class_id)
    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(verification_status=status)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Records"
    
    # Headers
    headers = ['Student Name', 'Registration Number', 'Class', 'Date/Time', 'Method', 'Status', 'Verified By']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="696cff", end_color="696cff", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, attendance in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=attendance.student.person.get_full_name())
        ws.cell(row=row, column=2, value=attendance.student.student_reg_number)
        ws.cell(row=row, column=3, value=attendance.class_obj.name)
        ws.cell(row=row, column=4, value=attendance.scan_time.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row, column=5, value=attendance.get_verification_method_display())
        ws.cell(row=row, column=6, value="Present" if attendance.verification_status == "success" else attendance.verification_status)
        ws.cell(row=row, column=7, value=attendance.recorded_by.get_full_name() if attendance.recorded_by else "")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_records.xlsx"'
    wb.save(response)
    return response

@login_required
def download_attendance_template(request):
    """Download Excel template for attendance import"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Template"
    
    # Headers
    headers = ['Student Registration Number', 'Class Code', 'Scan Time (Optional)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="696cff", end_color="696cff", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Example data
    ws.cell(row=2, column=1, value="S12345")
    ws.cell(row=2, column=2, value="CS101")
    ws.cell(row=2, column=3, value="2024-01-15 09:30:00")
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_import_template.xlsx"'
    wb.save(response)
    return response

@login_required
@csrf_exempt
def import_attendance_excel(request):
    """Import attendance records from Excel"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            from apps.classroom.services import AttendanceImportService
            excel_file = request.FILES['excel_file']
            service = AttendanceImportService()
            result = service.import_from_excel(excel_file, request.user)
            
            if result['success']:
                messages.success(request, f"Successfully imported {result['imported']} records. {result.get('skipped', 0)} skipped.")
            else:
                messages.error(request, result.get('error', 'Import failed'))
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
    else:
        messages.error(request, "No file selected")
    
    return redirect('classroom:list')

# apps/classroom/views.py - Add this method

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from django.urls import reverse_lazy
from django.http import HttpResponse
from .services import AttendanceImportService, AttendanceExportService


class AttendanceImportView(LoginRequiredMixin, TemplateView):
    """
    View for importing attendance records from Excel
    """
    template_name = 'classroom/attendance_import.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Import Attendance Records'
        return context
    
    def post(self, request):
        """Handle Excel file upload and import"""
        if request.method == 'POST' and request.FILES.get('excel_file'):
            try:
                excel_file = request.FILES['excel_file']
                
                # Validate file type
                if not excel_file.name.endswith(('.xlsx', '.xls', '.csv')):
                    messages.error(request, 'Please upload an Excel file (.xlsx, .xls) or CSV file')
                    return redirect('classroom:attendance_import')
                
                service = AttendanceImportService()
                result = service.import_from_excel(excel_file, request.user)
                
                if result['success']:
                    success_msg = f"Successfully imported {result['imported']} records."
                    if result['skipped'] > 0:
                        success_msg += f" Skipped {result['skipped']} records."
                    messages.success(request, success_msg)
                    
                    if result['errors']:
                        for error in result['errors'][:5]:
                            messages.warning(request, error)
                else:
                    messages.error(request, result.get('error', 'Import failed'))
                    
            except Exception as e:
                messages.error(request, f"Import failed: {str(e)}")
        else:
            messages.error(request, "No file selected")
        
        return redirect('classroom:attendance_list')


def download_attendance_template(request):
    """
    Download Excel template for attendance import
    """
    service = AttendanceImportService()
    output = service.download_template()
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="attendance_import_template.xlsx"'
    
    return response


def export_attendance_records(request):
    """
    Export attendance records to Excel
    """
    from apps.classroom.models import ClassAttendance
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    class_id = request.GET.get('class_id')
    
    queryset = ClassAttendance.objects.select_related('student__person', 'class_obj__academic_unit')
    
    if start_date:
        queryset = queryset.filter(scan_time__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(scan_time__date__lte=end_date)
    if class_id:
        queryset = queryset.filter(class_obj_id=class_id)
    
    service = AttendanceExportService()
    output = service.export_to_excel(queryset)
    
    filename = f"attendance_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response