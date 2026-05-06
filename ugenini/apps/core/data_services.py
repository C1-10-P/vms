import pandas as pd
import csv
import io
import json
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import transaction
from django.db.models import Count, Avg, Q
from django.core.cache import cache
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import logging
from django.utils import timezone
from datetime import timedelta
from django.db import models

logger = logging.getLogger(__name__)


class DataExportService:
    """
    Service for exporting data to various formats
    """
    
    @staticmethod
    def export_to_excel(data, sheet_name="Sheet1"):
        """Export data to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add headers
        if data and len(data) > 0:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for row_idx, row in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    value = row.get(key, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_to_csv(data):
        """Export data to CSV format"""
        output = io.StringIO()
        
        if data and len(data) > 0:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return output.getvalue()
    
    @staticmethod
    def export_to_pdf(title, data, columns):
        """Export data to PDF format"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        story = []
        
        # Title
        story.append(Paragraph(title, styles['Title']))
        story.append(Spacer(1, 12))
        
        # Table data
        table_data = [columns]
        for row in data:
            table_data.append([row.get(col, '') for col in columns])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        
        buffer.seek(0)
        return buffer


class DataImportService:
    """
    Service for importing data from various formats
    """
    
    @staticmethod
    def import_from_csv(file_content, model, field_mapping):
        """Import data from CSV file"""
        decoded = file_content.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(decoded))
        
        created_count = 0
        errors = []
        
        with transaction.atomic():
            for row_num, row in enumerate(csv_reader, 2):
                try:
                    # Map CSV columns to model fields
                    data = {}
                    for csv_field, model_field in field_mapping.items():
                        if csv_field in row:
                            data[model_field] = row[csv_field]
                    
                    # Create model instance
                    model.objects.create(**data)
                    created_count += 1
                    
                except Exception as e:
                    errors.append({
                        'row': row_num,
                        'error': str(e)
                    })
        
        return {
            'success': created_count,
            'errors': errors,
            'total': created_count + len(errors)
        }
    
    @staticmethod
    def import_from_excel(file, model, field_mapping):
        """Import data from Excel file"""
        df = pd.read_excel(file)
        
        created_count = 0
        errors = []
        
        with transaction.atomic():
            for idx, row in df.iterrows():
                try:
                    data = {}
                    for excel_field, model_field in field_mapping.items():
                        if excel_field in row:
                            value = row[excel_field]
                            if pd.isna(value):
                                value = None
                            data[model_field] = value
                    
                    model.objects.create(**data)
                    created_count += 1
                    
                except Exception as e:
                    errors.append({
                        'row': idx + 2,
                        'error': str(e)
                    })
        
        return {
            'success': created_count,
            'errors': errors,
            'total': created_count + len(errors)
        }


class DataAggregationService:
    """
    Service for data aggregation and analytics
    """
    
    @staticmethod
    def get_attendance_analytics(start_date, end_date, group_by='day'):
        """Get attendance analytics for date range"""
        from apps.classroom.models import ClassAttendance
        from django.db.models import Count, Avg, Q
        from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
        
        queryset = ClassAttendance.objects.filter(
            scan_time__date__gte=start_date,
            scan_time__date__lte=end_date,
            verification_status='success'
        )
        
        if group_by == 'day':
            trunc_func = TruncDate('scan_time')
        elif group_by == 'week':
            trunc_func = TruncWeek('scan_time')
        else:
            trunc_func = TruncMonth('scan_time')
        
        analytics = queryset.annotate(
            period=trunc_func
        ).values('period').annotate(
            total=Count('id'),
            unique_students=Count('student', distinct=True),
            avg_confidence=Avg('confidence_score')
        ).order_by('period')
        
        return list(analytics)
    
    @staticmethod
    def get_visitor_analytics(start_date, end_date):
        """Get visitor analytics for date range"""
        from apps.vms.models import VisitorVisit, VisitorMovement
        
        visits = VisitorVisit.objects.filter(
            check_in_time__date__gte=start_date,
            check_in_time__date__lte=end_date,
            status='completed'
        )
        
        # Average visit duration
        avg_duration = visits.extra(
            select={'duration': "EXTRACT(EPOCH FROM (check_out_time - check_in_time))"}
        ).aggregate(avg=models.Avg('duration'))
        
        # Popular zones
        popular_zones = VisitorMovement.objects.filter(
            timestamp__date__gte=start_date,
            timestamp__date__lte=end_date,
            event_type='enter'
        ).values('zone__name').annotate(
            visit_count=Count('visitor', distinct=True)
        ).order_by('-visit_count')[:10]
        
        return {
            'total_visits': visits.count(),
            'average_duration_seconds': avg_duration.get('avg', 0),
            'popular_zones': list(popular_zones),
            'daily_trend': DataAggregationService._get_daily_visitor_trend(start_date, end_date)
        }
    
    @staticmethod
    def _get_daily_visitor_trend(start_date, end_date):
        """Get daily visitor trend"""
        from apps.vms.models import VisitorVisit
        from django.db.models import Count
        from django.db.models.functions import TruncDate
        
        return list(VisitorVisit.objects.filter(
            check_in_time__date__gte=start_date,
            check_in_time__date__lte=end_date
        ).annotate(
            date=TruncDate('check_in_time')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date'))
    
    @staticmethod
    def get_device_analytics(days=30):
        """Get device analytics for last N days"""
        from apps.firmware.models import NodeHeartbeat, EdgeNode
        from django.db.models import Avg, Count
        
        cutoff = timezone.now() - timedelta(days=days)
        
        heartbeats = NodeHeartbeat.objects.filter(timestamp__gte=cutoff)
        
        return {
            'total_devices': EdgeNode.objects.count(),
            'avg_uptime': heartbeats.aggregate(Avg('uptime_seconds'))['uptime_seconds__avg'],
            'avg_battery': heartbeats.aggregate(Avg('battery_level'))['battery_level__avg'],
            'avg_rssi': heartbeats.aggregate(Avg('rssi'))['rssi__avg'],
            'uptime_by_device': list(heartbeats.values('node__name').annotate(
                avg_uptime=Avg('uptime_seconds')
            )),
            'low_battery_devices': EdgeNode.objects.filter(battery_level__lt=20).count()
        }
    
    @staticmethod
    def get_access_analytics(hours=24):
        """Get access control analytics for last N hours"""
        from apps.access.models import AccessLog
        from django.db.models import Count
        
        cutoff = timezone.now() - timedelta(hours=hours)
        logs = AccessLog.objects.filter(access_time__gte=cutoff)
        
        total = logs.count()
        granted = logs.filter(result='granted').count()
        denied = logs.filter(result='denied').count()
        
        return {
            'total_attempts': total,
            'granted': granted,
            'denied': denied,
            'success_rate': (granted / total * 100) if total > 0 else 0,
            'by_zone': list(logs.values('zone__name').annotate(
                attempts=Count('id'),
                granted=Count('id', filter=models.Q(result='granted'))
            )),
            'peak_hour': DataAggregationService._get_peak_access_hour(logs)
        }
    
    @staticmethod
    def _get_peak_access_hour(logs):
        """Get peak access hour"""
        from django.db.models.functions import ExtractHour
        peak = logs.annotate(
            hour=ExtractHour('access_time')
        ).values('hour').annotate(
            count=Count('id')
        ).order_by('-count').first()
        
        return peak['hour'] if peak else None


class CacheService:
    """
    Service for managing cached data
    """
    
    CACHE_KEYS = {
        'DASHBOARD_STATS': 'dashboard_stats',
        'INSTITUTIONS_LIST': 'institutions_list',
        'ACTIVE_VISITORS': 'active_visitors',
        'ONLINE_DEVICES': 'online_devices',
        'ATTENDANCE_TODAY': 'attendance_today'
    }
    
    @classmethod
    def get_or_set(cls, key, callback, timeout=300):
        """Get from cache or set from callback"""
        cached_value = cache.get(key)
        if cached_value is not None:
            return cached_value
        
        value = callback()
        cache.set(key, value, timeout)
        return value
    
    @classmethod
    def invalidate(cls, key):
        """Invalidate a cache key"""
        cache.delete(key)
    
    @classmethod
    def invalidate_group(cls, pattern):
        """Invalidate all keys matching pattern"""
        cache.delete_pattern(pattern)
    
    @classmethod
    def refresh_dashboard_stats(cls):
        """Refresh dashboard statistics cache"""
        cache.delete(cls.CACHE_KEYS['DASHBOARD_STATS'])
    
    @classmethod
    def get_dashboard_stats(cls, callback):
        """Get cached dashboard stats"""
        return cls.get_or_set(cls.CACHE_KEYS['DASHBOARD_STATS'], callback, timeout=60)


class ReportGenerationService:
    """
    Service for generating various reports
    """
    
    @staticmethod
    def generate_attendance_report(start_date, end_date, department_id=None, class_id=None):
        """Generate comprehensive attendance report"""
        from apps.classroom.models import ClassAttendance
        from django.db.models import Count, Q
        
        queryset = ClassAttendance.objects.filter(
            scan_time__date__gte=start_date,
            scan_time__date__lte=end_date
        )
        
        if department_id:
            queryset = queryset.filter(student__department_id=department_id)
        if class_id:
            queryset = queryset.filter(class_obj_id=class_id)
        
        # Summary statistics
        total_records = queryset.count()
        successful = queryset.filter(verification_status='success').count()
        failed = queryset.filter(verification_status='failed').count()
        
        # Top performing students
        top_students = queryset.filter(
            verification_status='success'
        ).values('student__person__first_name', 'student__person__last_name',
                 'student__student_reg_number').annotate(
            attendance_count=Count('id')
        ).order_by('-attendance_count')[:10]
        
        # Daily breakdown
        daily_breakdown = queryset.extra(
            {'date': "DATE(scan_time)"}
        ).values('date').annotate(
            count=Count('id'),
            successful=Count('id', filter=Q(verification_status='success')),
            failed=Count('id', filter=Q(verification_status='failed'))
        ).order_by('date')
        
        # Hourly breakdown
        from django.db.models.functions import ExtractHour
        hourly_breakdown = queryset.annotate(
            hour=ExtractHour('scan_time')
        ).values('hour').annotate(
            count=Count('id')
        ).order_by('hour')
        
        return {
            'summary': {
                'start_date': start_date,
                'end_date': end_date,
                'total_records': total_records,
                'successful': successful,
                'failed': failed,
                'success_rate': (successful / total_records * 100) if total_records > 0 else 0
            },
            'top_students': list(top_students),
            'daily_breakdown': list(daily_breakdown),
            'hourly_breakdown': list(hourly_breakdown)
        }
    
    @staticmethod
    def generate_visitor_report(start_date, end_date):
        """Generate visitor report"""
        from apps.vms.models import VisitorVisit
        
        visits = VisitorVisit.objects.filter(
            check_in_time__date__gte=start_date,
            check_in_time__date__lte=end_date
        )
        
        # Visitor statistics
        total_visitors = visits.values('visitor').distinct().count()
        total_visits = visits.count()
        average_duration = visits.extra(
            select={'duration': "EXTRACT(EPOCH FROM (check_out_time - check_in_time))"}
        ).aggregate(avg=models.Avg('duration'))
        
        # Top organizations
        top_orgs = visits.values('visitor__organization').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        # Visit purposes
        purposes = visits.values('visitor__purpose').annotate(
            count=Count('id')
        )
        
        # Daily breakdown
        from django.db.models.functions import TruncDate
        daily = visits.annotate(
            date=TruncDate('check_in_time')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
        
        return {
            'summary': {
                'start_date': start_date,
                'end_date': end_date,
                'total_visitors': total_visitors,
                'total_visits': total_visits,
                'average_duration_seconds': average_duration.get('avg', 0)
            },
            'top_organizations': list(top_orgs),
            'purposes': list(purposes),
            'daily_breakdown': list(daily)
        }